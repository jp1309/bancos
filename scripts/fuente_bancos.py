#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Validaciones estructurales de los boletines bancarios descargados."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import zipfile

import pandas as pd
from openpyxl import load_workbook


HOJAS_REQUERIDAS = ("BAL", "PYG", "CAMEL")


@dataclass(frozen=True)
class InspeccionExcel:
    ruta: Path
    fechas_maximas: dict[str, pd.Timestamp]

    @property
    def fecha_corte(self) -> pd.Timestamp:
        return max(self.fechas_maximas.values())


def _fechas_fila_cinco(hoja) -> list[pd.Timestamp]:
    fila = next(hoja.iter_rows(min_row=5, max_row=5, values_only=True))
    fechas: list[pd.Timestamp] = []
    for valor in fila:
        if isinstance(valor, (datetime, pd.Timestamp)):
            fechas.append(pd.Timestamp(valor).normalize() + pd.offsets.MonthEnd(0))
    return fechas


def inspeccionar_excel_banco(ruta: Path) -> InspeccionExcel:
    """Exige BAL/PYG/CAMEL y obtiene la fecha maxima real de cada hoja."""
    ruta = Path(ruta)
    if ruta.suffix.lower() != ".xlsx":
        raise ValueError(f"Formato no soportado para validacion: {ruta.name}")

    try:
        libro = load_workbook(ruta, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Excel invalido {ruta.name}: {exc}") from exc

    try:
        faltantes = [hoja for hoja in HOJAS_REQUERIDAS if hoja not in libro.sheetnames]
        if faltantes:
            raise ValueError(
                f"{ruta.name} no contiene hojas requeridas: {', '.join(faltantes)}"
            )

        fechas_maximas: dict[str, pd.Timestamp] = {}
        for nombre in HOJAS_REQUERIDAS:
            fechas = _fechas_fila_cinco(libro[nombre])
            if not fechas:
                raise ValueError(f"{ruta.name}/{nombre} no contiene fechas en la fila 5")
            fechas_maximas[nombre] = max(fechas)
    finally:
        libro.close()

    fechas_unicas = set(fechas_maximas.values())
    if len(fechas_unicas) != 1:
        detalle = ", ".join(
            f"{hoja}={fecha:%Y-%m-%d}" for hoja, fecha in fechas_maximas.items()
        )
        raise ValueError(f"{ruta.name} mezcla fechas de corte: {detalle}")

    return InspeccionExcel(ruta=ruta, fechas_maximas=fechas_maximas)


def validar_directorio_fuentes(
    directorio: Path,
    fecha_esperada: pd.Timestamp,
    bancos_esperados: int,
) -> list[InspeccionExcel]:
    """Valida cantidad, estructura y fecha uniforme de todos los bancos."""
    directorio = Path(directorio)
    archivos = sorted(directorio.glob("**/*.xlsx"))
    if len(archivos) != bancos_esperados:
        raise ValueError(
            f"Se esperaban {bancos_esperados} Excel y se encontraron {len(archivos)}"
        )

    inspecciones = [inspeccionar_excel_banco(ruta) for ruta in archivos]
    fechas = {item.fecha_corte for item in inspecciones}
    if len(fechas) != 1:
        detalle = ", ".join(
            f"{item.ruta.parent.name}={item.fecha_corte:%Y-%m-%d}"
            for item in inspecciones
        )
        raise ValueError(f"Las entidades no tienen una fecha de corte uniforme: {detalle}")

    fecha_fuente = fechas.pop()
    fecha_esperada = pd.Timestamp(fecha_esperada).normalize() + pd.offsets.MonthEnd(0)
    if fecha_fuente > fecha_esperada:
        raise ValueError(
            f"La fuente contiene {fecha_fuente:%Y-%m-%d}, posterior al objetivo "
            f"{fecha_esperada:%Y-%m-%d}"
        )
    return inspecciones


def validar_zip(ruta: Path) -> list[str]:
    """Comprueba que el ZIP sea integro, seguro y contenga un unico XLSX."""
    ruta = Path(ruta)
    try:
        with zipfile.ZipFile(ruta) as archivo:
            corrupto = archivo.testzip()
            if corrupto:
                raise ValueError(f"ZIP corrupto {ruta.name}: {corrupto}")
            nombres = [info.filename for info in archivo.infolist() if not info.is_dir()]
    except zipfile.BadZipFile as exc:
        raise ValueError(f"ZIP invalido: {ruta.name}") from exc

    for nombre in nombres:
        partes = Path(nombre.replace("\\", "/")).parts
        if Path(nombre).is_absolute() or ".." in partes:
            raise ValueError(f"Ruta insegura dentro de {ruta.name}: {nombre}")

    excels = [nombre for nombre in nombres if nombre.lower().endswith(".xlsx")]
    if len(excels) != 1:
        raise ValueError(
            f"{ruta.name} debe contener un unico XLSX; contiene {len(excels)}"
        )
    return excels
