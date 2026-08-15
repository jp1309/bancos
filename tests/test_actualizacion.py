import json
import runpy
import tempfile
import unittest
from pathlib import Path
import sys
import zipfile

import pandas as pd
from openpyxl import Workbook


ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import fuente_bancos
import nombres_bancos
import portal_bancos
import validar_actualizacion as validador


def crear_excel(ruta: Path, fecha: str, omitir: str | None = None):
    libro = Workbook()
    libro.remove(libro.active)
    for nombre in fuente_bancos.HOJAS_REQUERIDAS:
        if nombre == omitir:
            continue
        hoja = libro.create_sheet(nombre)
        hoja.cell(5, 3, pd.Timestamp(fecha).to_pydatetime())
    libro.save(ruta)


class FuenteBancosTests(unittest.TestCase):
    def test_inspeccion_exige_tres_hojas_y_fecha_uniforme(self):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "Banco.xlsx"
            crear_excel(ruta, "2026-06-30")
            inspeccion = fuente_bancos.inspeccionar_excel_banco(ruta)
            self.assertEqual(inspeccion.fecha_corte, pd.Timestamp("2026-06-30"))
            self.assertEqual(set(inspeccion.fechas_maximas), {"BAL", "PYG", "CAMEL"})

    def test_inspeccion_rechaza_hoja_faltante(self):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "Banco.xlsx"
            crear_excel(ruta, "2026-06-30", omitir="PYG")
            with self.assertRaisesRegex(ValueError, "PYG"):
                fuente_bancos.inspeccionar_excel_banco(ruta)

    def test_directorio_rechaza_fechas_mezcladas(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            crear_excel(base / "A.xlsx", "2026-06-30")
            crear_excel(base / "B.xlsx", "2026-05-31")
            with self.assertRaisesRegex(ValueError, "fecha de corte uniforme"):
                fuente_bancos.validar_directorio_fuentes(
                    base, pd.Timestamp("2026-06-30"), bancos_esperados=2
                )

    def test_zip_debe_contener_un_unico_xlsx(self):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "fuente.zip"
            with zipfile.ZipFile(ruta, "w") as archivo:
                archivo.writestr("A.xlsx", b"uno")
                archivo.writestr("B.xlsx", b"dos")
            with self.assertRaisesRegex(ValueError, "unico XLSX"):
                fuente_bancos.validar_zip(ruta)


class PortalBancosTests(unittest.TestCase):
    def setUp(self):
        self.archivos = [
            {
                "nombre": f"Series Banco {i} JULIO 2026.zip",
                "id": f"id-{i}",
                "url": (
                    "https://www.superbancos.gob.ec/estadisticas/portalestudios/"
                    "wp-admin/admin-ajax.php?action=shareonedrive-download"
                    f"&id=id-{i}&listtoken=token-vigente"
                ),
            }
            for i in range(2)
        ]

    def test_usa_enlace_dinamico_del_dom(self):
        preparados = portal_bancos.preparar_archivos_dom(self.archivos)
        self.assertIn("listtoken=token-vigente", preparados[0]["url"])

    def test_rechaza_host_ajeno(self):
        self.archivos[0]["url"] = "https://example.com/archivo.zip?id=id-0"
        with self.assertRaisesRegex(ValueError, "Host de descarga no permitido"):
            portal_bancos.preparar_archivos_dom(self.archivos)

    def test_detecta_fuente_rezagada_sin_descargar(self):
        archivos = [
            dict(a, nombre=a["nombre"].replace("JULIO", "JUNIO"))
            for a in self.archivos
        ]
        self.assertEqual(
            portal_bancos.clasificar_publicacion(archivos, "JULIO 2026", 2),
            "rezagada",
        )

    def test_rechaza_publicacion_parcial(self):
        self.archivos[0]["nombre"] = self.archivos[0]["nombre"].replace(
            "JULIO", "JUNIO"
        )
        with self.assertRaisesRegex(ValueError, "Publicacion parcial"):
            portal_bancos.clasificar_publicacion(
                self.archivos, "JULIO 2026", 2
            )


class NombresBancosTests(unittest.TestCase):
    def test_unifica_acentos_y_mayusculas_del_portal(self):
        self.assertEqual(nombres_bancos.normalizar_banco("Pacifico"), "Pacífico")
        self.assertEqual(
            nombres_bancos.normalizar_banco("Comercial de manabí"),
            "Comercial de Manabí",
        )
        self.assertEqual(
            nombres_bancos.normalizar_banco("General Rumin\u0303ahui"),
            "General Rumiñahui",
        )

    def test_todos_los_bancos_canonicos_tienen_color_en_la_ui(self):
        configuracion_ui = runpy.run_path(ROOT / "config" / "indicator_mapping.py")
        self.assertTrue(
            set(nombres_bancos.NOMBRES_CANONICOS).issubset(
                configuracion_ui["COLORES_BANCOS"]
            )
        )


class ValidadorPublicacionTests(unittest.TestCase):
    def setUp(self):
        self.temporal = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporal.cleanup)
        self.master = Path(self.temporal.name)
        self.master_original = validador.MASTER_DATA_DIR
        validador.MASTER_DATA_DIR = self.master
        self.addCleanup(setattr, validador, "MASTER_DATA_DIR", self.master_original)

    def escribir_datasets(self, pyg_rezagado=False):
        bancos = ["A", "B"]
        fechas = pd.to_datetime(["2026-05-31", "2026-06-30"])

        balance = pd.DataFrame([
            {"banco": banco, "fecha": fecha, "codigo": "1", "cuenta": "ACTIVO",
             "valor": 1.0, "nivel": 1}
            for banco in bancos for fecha in fechas
        ])
        pyg = pd.DataFrame([
            {"banco": banco, "fecha": fecha, "codigo": "GDE", "cuenta": "GANANCIA",
             "valor_acumulado": 1.0, "valor_mes": 1.0, "valor_12m": 1.0}
            for banco in bancos
            for fecha in fechas
            if not (pyg_rezagado and banco == "B" and fecha == fechas[-1])
        ])
        camel = pd.DataFrame([
            {"banco": banco, "fecha": fecha, "codigo": "ROE", "indicador": "ROE",
             "valor": 0.1, "categoria": "E"}
            for banco in bancos for fecha in fechas
        ])
        balance.to_parquet(self.master / "balance.parquet", index=False)
        pyg.to_parquet(self.master / "pyg.parquet", index=False)
        camel.to_parquet(self.master / "camel.parquet", index=False)
        (self.master / "metadata.json").write_text(json.dumps({
            "fecha_max": "2026-06-30 00:00:00",
            "bancos_error": [],
        }), encoding="utf-8")

    def test_valida_tres_datasets_completos(self):
        self.escribir_datasets()
        estado = validador.validar_actualizacion(
            pd.Timestamp("2026-06-30"), bancos_esperados=2
        )
        self.assertEqual(estado["pyg"]["bancos_ultimo_mes"], ["A", "B"])

    def test_rechaza_banco_rezagado_en_pyg(self):
        self.escribir_datasets(pyg_rezagado=True)
        with self.assertRaisesRegex(ValueError, "pyg: solo 1 bancos"):
            validador.validar_actualizacion(
                pd.Timestamp("2026-06-30"), bancos_esperados=2
            )

    def test_rechaza_perdida_de_historia(self):
        self.escribir_datasets()
        anterior = validador.capturar_estado()
        anterior["balance"]["meses"].insert(0, "2026-04")
        with self.assertRaisesRegex(ValueError, "perdio meses publicados"):
            validador.validar_actualizacion(
                pd.Timestamp("2026-06-30"), anterior, bancos_esperados=2
            )


if __name__ == "__main__":
    unittest.main()
